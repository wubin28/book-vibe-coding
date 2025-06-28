import openpyxl
import sys
from operator import itemgetter

def analyze_receivables(file_path):
    try:
        # 加载工作簿和工作表
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # 获取目标部门列表（表头）
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        target_departments = [cell for cell in header_row[2:] if cell is not None]
        
        # 收集所有应收款记录
        receivables = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            creditor = row[0]  # 债权部门
            account_type = row[1]  # 应收/应付类型
            
            if account_type != "应收款":
                continue
                
            for idx, amount in enumerate(row[2:2+len(target_departments)]):
                # 跳过无效数据：空值、0值、相同部门
                if amount is None or amount == "" or amount == 0:
                    continue
                if creditor == target_departments[idx]:
                    continue
                    
                debtor = target_departments[idx]  # 目标部门
                receivables.append((creditor, debtor, amount))
        
        # 按金额降序排序并取前三
        receivables.sort(key=itemgetter(2), reverse=True)
        top_three = receivables[:3]
        
        # 格式化输出结果
        for i, (creditor, debtor, amount) in enumerate(top_three, 1):
            print(f"{i}. {creditor}应收{debtor}：{amount}元")
            
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    file_name = "部门对账表-脱敏-样例.xlsx"
    if len(sys.argv) > 1:
        file_name = sys.argv[1]
    
    analyze_receivables(file_name)